import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
import io
import os
from glob import glob
import pydeck as pdk
import folium
from streamlit_folium import st_folium


# ----------------- Utility Functions -----------------
@st.cache_data
def load_excel(file_bytes):
    return load_workbook(io.BytesIO(file_bytes))

def preprocess_sheet(sheet):
    for merged_range in list(sheet.merged_cells.ranges):
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = value

    for col in reversed(range(1, sheet.max_column + 1)):
        if all(sheet.cell(row=row, column=col).value in [None, ""] for row in range(1, sheet.max_row + 1)):
            sheet.delete_cols(col)

    empty_count = 0
    for row in range(1, sheet.max_row + 1):
        is_empty = all(sheet.cell(row=row, column=col).value in [None, ""] for col in range(1, sheet.max_column + 1))
        if is_empty:
            empty_count += 1
            if empty_count == 2:
                sheet.delete_rows(row, sheet.max_row - row + 1)
                break
        else:
            empty_count = 0

def load_workbook_with_images(file_like):
    wb = load_workbook(file_like)
    images_per_sheet = {}
    for sheet in wb.worksheets:
        if sheet.title in ["Audit Data", "Etilt"]:
            continue
        if hasattr(sheet, "_images") and sheet._images:
            images_per_sheet[sheet.title] = sheet._images
    return wb, images_per_sheet

def apply_image_rotation(img_obj, pil_image, global_rotation=0):
    try:
        if hasattr(img_obj, 'rotation') and img_obj.rotation:
            pil_image = pil_image.rotate(-img_obj.rotation, expand=True)
    except Exception:
        pass
    if global_rotation:
        pil_image = pil_image.rotate(global_rotation, expand=True)
    return pil_image


# ----------------- Extraction Functions -----------------
def extract_site_configuration(sheet):
    columns = {
        "site_id": None, "site_type": None, "ant_ht": None, "str_ht": None,
        "latitude": None, "longitude": None
    }

    for col in range(1, sheet.max_column + 1):
        for row in range(1, 4):
            val = sheet.cell(row=row, column=col).value
            if not val:
                continue
            val_lower = str(val).strip().lower()
            if "site id" in val_lower:
                columns["site_id"] = col
            elif "site type" in val_lower:
                columns["site_type"] = col
            elif "ant" in val_lower and "ht" in val_lower:
                columns["ant_ht"] = col
            elif ("structure" in val_lower or "struc" in val_lower) and "ht" in val_lower:
                columns["str_ht"] = col
            elif "lat" in val_lower:
                columns["latitude"] = col
            elif "lon" in val_lower or "lng" in val_lower:
                columns["longitude"] = col

    for row in range(4, sheet.max_row + 1):
        site_name = sheet.cell(row=row, column=columns["site_id"]).value if columns["site_id"] else None
        site_type = sheet.cell(row=row, column=columns["site_type"]).value if columns["site_type"] else None
        ant_ht = sheet.cell(row=row, column=columns["ant_ht"]).value if columns["ant_ht"] else None
        str_ht = sheet.cell(row=row, column=columns["str_ht"]).value if columns["str_ht"] else None
        lat = sheet.cell(row=row, column=columns["latitude"]).value if columns["latitude"] else None
        lon = sheet.cell(row=row, column=columns["longitude"]).value if columns["longitude"] else None

        if any([site_name, site_type, ant_ht, str_ht]):
            return {
                "SiteName": str(site_name).strip() if site_name else "",
                "SiteType": str(site_type).strip() if site_type else "",
                "AntHT": str(ant_ht).strip() if ant_ht else "",
                "StrHT": str(str_ht).strip() if str_ht else "",
                "Latitude": float(lat) if lat else None,
                "Longitude": float(lon) if lon else None
            }
    return None


def extract_azimuths_and_tilts(sheet):
    result = {"2G": {"Azimuth": [], "Tilt": []}, "3G": {"Azimuth": [], "Tilt": []}, "4G": {"Azimuth": [], "Tilt": []}, "5G": {"Antenna 1": {"Azimuth": [], "Tilt": []}, "Antenna 2": {"Azimuth": [], "Tilt": []}}}
    for col in range(1, sheet.max_column + 1):
        tech = sheet.cell(row=2, column=col).value
        if not tech:
            continue
        tech = str(tech).strip().upper()
        label = sheet.cell(row=3, column=col).value
        label = str(label).strip().lower() if label else ""
        values = [sheet.cell(row=row, column=col).value for row in range(4, sheet.max_row + 1) if sheet.cell(row=row, column=col).value not in (None, "")]
        if tech == "5G":
            if "ant-1" in label or "azimuth-1" in label or "azimuth" in label or "azimut" in label:
                result["5G"]["Antenna 1"]["Azimuth"].extend(values)
            elif "ant-2" in label or "azimuth-2" in label:
                result["5G"]["Antenna 2"]["Azimuth"].extend(values)
            elif "mech-1" in label or "mtilt-1" in label or "tilt-1" in label:
                result["5G"]["Antenna 1"]["Tilt"].extend(values)
            elif "mech-2" in label or "mtilt-2" in label or "tilt-2" in label:
                result["5G"]["Antenna 2"]["Tilt"].extend(values)
        elif tech in ["2G", "3G", "4G"]:
            if "tilt" in label or "mech" in label or "mtilt" in label:
                result[tech]["Tilt"].extend(values)
            elif "azimuth" in label or "ant" in label:
                result[tech]["Azimuth"].extend(values)
    return result

def extract_antenna_types(sheet):
    result = {"2G": [], "3G": [], "4G": [], "5G": {"Antenna 1": [], "Antenna 2": []}}
    for col in range(1, sheet.max_column + 1):
        header1 = sheet.cell(row=1, column=col).value
        if not header1 or 'antenna' not in str(header1).lower():
            continue
        tech = sheet.cell(row=2, column=col).value
        if not tech:
            continue
        tech = str(tech).strip().upper()
        values = [str(sheet.cell(row=row, column=col).value).strip() for row in range(4, sheet.max_row + 1) if sheet.cell(row=row, column=col).value not in (None, "")]
        if tech == "5G-1":
            result["5G"]["Antenna 1"].extend(values)
        elif tech == "5G-2":
            result["5G"]["Antenna 2"].extend(values)
        elif tech in result:
            result[tech].extend(values)
    return result

# ----------------- App Layout -----------------
st.set_page_config(layout="wide", page_title="RF Site Audit Dashboard v4")

# Load logo
logo = Image.open("your_logo.png")  # Make sure this image is in your working directory

# Title & subtitle
st.markdown("""
    <div style='text-align: center; padding-top: 1px;'>
        <h1 style='font-size: 36px; color: #003366;'>📡 RF Site Audit Dashboard</h1>
        <h4 style='color: #666;'>Analyze Site Engineering Parameters & 360 View</h4>
        <hr style='border:1px solid #ccc; width: 60%; margin: auto;'/>
    </div>
""", unsafe_allow_html=True)

#SideBar
with st.sidebar:
    st.image(logo, width=100)
    st.markdown("---")  # Optional horizontal line
    st.header("🔧 Controls")
    st.markdown("---")
    rotation_angle = st.selectbox("🌀 Rotate All Images", [0, 90, 180, 270], index=0)
    folder_path = st.text_input("📁 Folder Path")
    search_query = st.text_input("🔍 Search by Site Name")

excel_files = glob(os.path.join(folder_path, "*.xlsx")) if folder_path else []

# Filter files based on search query
filtered_files = [f for f in excel_files if search_query.lower() in os.path.basename(f).lower()] if search_query else excel_files

# Extract file names only for display
file_names = [os.path.basename(f) for f in filtered_files]

# Let user select file by name
selected_file_name = st.sidebar.selectbox("📄 Select Audit File", file_names) if file_names else None

# Get full path for the selected file
selected_file = None
if selected_file_name:
    selected_file = next((f for f in filtered_files if os.path.basename(f) == selected_file_name), None)


if not selected_file:
    st.warning("Please provide a valid folder path with .xlsx files.")
    st.stop()

with open(selected_file, "rb") as f:
    file_bytes = f.read()

wb = load_excel(file_bytes)
sheet = next((ws for ws in wb.worksheets if ws.title.lower() == "audit data"), None)
if not sheet:
    st.error("No 'Audit Data' sheet found.")
    st.stop()

preprocess_sheet(sheet)
site_info = extract_site_configuration(sheet)
az_tilt = extract_azimuths_and_tilts(sheet)
ant_types = extract_antenna_types(sheet)

if site_info.get("Latitude") and site_info.get("Longitude"):
    st.subheader("🗼 Site Details and Location")

    # Create two columns: Left (site info), Right (map)
    col1, col2 = st.columns([1, 1.5])  # Adjust widths as needed

    with col1:
        st.markdown("### 📝 Site Overview")
        st.metric("Site Name", site_info["SiteName"])
        st.metric("Site Type", site_info["SiteType"])
        st.metric("Antenna Height", site_info["AntHT"])
        st.metric("Structure Height", site_info["StrHT"])
        #st.metric("Latitude", f"{site_info['Latitude']:.6f}")
        #st.metric("Longitude", f"{site_info['Longitude']:.6f}")

    with col2:
        st.markdown("### 🗺️ Location Map")
        import folium
        from streamlit_folium import st_folium

        m = folium.Map(
            location=[site_info["Latitude"], site_info["Longitude"]],
            zoom_start=16,
            tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
            attr='Esri Satellite'
        )
        folium.Marker(
            [site_info["Latitude"], site_info["Longitude"]],
            popup=f"<b>{site_info['SiteName']}</b>",
            tooltip="Click for site info",
            icon=folium.Icon(color="red", icon="info-sign")
        ).add_to(m)

        # Show smaller map size
        st_folium(m, width=500, height=400)




st.markdown("---")
# Format azimuth/tilt/antenna table
def create_sector_df(tech, az, tilt, antennas):
    rows = []
    max_len = max(len(az), len(tilt), len(antennas))
    for i in range(max_len):
        rows.append({
            "Technology": tech,
            "Sector": chr(65 + i),
            "Azimuth": az[i] if i < len(az) else "",
            "Tilt": tilt[i] if i < len(tilt) else "",
            "Antenna Type": antennas[i] if i < len(antennas) else ""
        })
    return rows

all_data = []
for tech in ["2G", "3G", "4G"]:
    all_data.extend(create_sector_df(tech, az_tilt[tech]["Azimuth"], az_tilt[tech]["Tilt"], ant_types[tech]))
all_data.extend(create_sector_df("5G Antenna 1", az_tilt["5G"]["Antenna 1"]["Azimuth"], az_tilt["5G"]["Antenna 1"]["Tilt"], ant_types["5G"]["Antenna 1"]))
all_data.extend(create_sector_df("5G Antenna 2", az_tilt["5G"]["Antenna 2"]["Azimuth"], az_tilt["5G"]["Antenna 2"]["Tilt"], ant_types["5G"]["Antenna 2"]))

st.subheader("📊 Sector Data Table")
df = pd.DataFrame(all_data)
st.dataframe(df, use_container_width=True)

# ----------------- Image Display -----------------
wb, image_dict = load_workbook_with_images(io.BytesIO(file_bytes))

if image_dict:
    selected_tab = st.selectbox("🖼️ Select Image Tab", list(image_dict.keys()))
    if selected_tab in image_dict:
        images_with_pos = []
        for img in image_dict[selected_tab]:
            anchor = img.anchor
            row = getattr(anchor, "_from", anchor).row + 1
            col = getattr(anchor, "_from", anchor).col + 1
            images_with_pos.append((img, row, col))

        if selected_tab.lower() == "panaromics":
            st.subheader("🌄 Panoramic Images")
            sorted_images = sorted(images_with_pos, key=lambda x: (x[1], x[2]))
            cols = st.columns(3)
            for idx, (img, row, col) in enumerate(sorted_images):
                image_bytes = io.BytesIO(img._data())
                rotated_image = apply_image_rotation(img, Image.open(image_bytes), global_rotation=rotation_angle)
                degree = idx * 30
                caption = f"Panoramic {degree}°"
                with cols[idx % 3]:
                    st.markdown(f"**{caption}**", help=f"R{row} C{col}")
                    st.image(rotated_image, use_container_width=True)

        else:
            # Grouping logic
            grouped_images = []
            current_group_start = None
            group_number = 0

            for img, row, col in sorted(images_with_pos, key=lambda x: (x[1], x[2])):
                if current_group_start is None or abs(row - current_group_start) > 6:
                    current_group_start = row
                    group_number += 1
                grouped_images.append((img, group_number, col, row))

            used_group_numbers = sorted(set(group for _, group, _, _ in grouped_images))

            label_sets = {
                "sector view": ["From Antenna Backside", "Sector View", "Antenna Front Side", "5G Antenna Backside", "5G Sector View"],
                "m-tilt": [f"Antenna-{i+1} M-Tilt" for i in range(len(used_group_numbers))],
                "antenna": [f"Antenna-{i+1}" for i in range(len(used_group_numbers))]
            }

            group_labels = label_sets.get(selected_tab.lower(), [f"Group {i+1}" for i in range(len(used_group_numbers))])

            for i, group_number in enumerate(used_group_numbers):
                label = group_labels[i] if i < len(group_labels) else f"View {group_number}"
                with st.expander(f"🔹 {label}", expanded=False):
                    group_images = [(img, col, row) for img, grp, col, row in grouped_images if grp == group_number]
                    group_images.sort(key=lambda x: x[1])
                    cols = st.columns(min(3, len(group_images)))
                    for idx, (img, col, row) in enumerate(group_images):
                        with cols[idx % len(cols)]:
                            image_bytes = io.BytesIO(img._data())
                            rotated_image = apply_image_rotation(img, Image.open(image_bytes), global_rotation=rotation_angle)
                            st.markdown(f"**{label} - Sector-{idx+1}**", help=f"R{row} C{col}")
                            st.image(rotated_image, use_container_width=True)


else:
    st.info("No image-containing sheets found in this report.")

# ----------------- Footer -----------------
st.markdown("""
---
<small>© 2025📡 | RF Audit Dashboard | ahsan.khan@du.ae </small>
""", unsafe_allow_html=True)
